"""
Update the ORIGINAL No_Website_Leads_Combined.xlsx in place:
- Add 'Demo URL' header in column N
- Fill each row with the live GitHub Pages URL as a clickable hyperlink
- Preserve all existing data, filters, frozen panes, and formatting
"""
import openpyxl
from openpyxl.styles import Font
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(__file__))

from generate_sites import (
    CATEGORY_ROUTING, CATEGORY_SLUG, city_info,
    slugify, clean_name,
)

XLSX = r'C:/Users/jmang/OneDrive/Desktop/No_Website_Leads_Combined.xlsx'
BASE = 'https://jaggyai.github.io/bc-demos'

# Column to place the URL in (N = 14)
URL_COL = 14

# Style for hyperlinked cells — standard "blue underline"
LINK_FONT = Font(color='0000EE', underline='single', name='Calibri', size=11)

def build_url(row_dict, used_slugs):
    """Mirror the generator's routing logic for a single row dict (header→value)."""
    cat1 = row_dict.get('Category 1')
    if not cat1:
        return None
    cat1_str = str(cat1).strip()
    cfg_key = CATEGORY_ROUTING.get(cat1_str)
    if cfg_key is None:
        return None
    cat_slug = CATEGORY_SLUG[cfg_key]

    raw_name = str(row_dict.get('Business Name', '')).strip()
    if not raw_name:
        return None
    name = clean_name(raw_name)
    biz_slug = slugify(name) or 'business'

    raw_city = row_dict.get('City')
    info = city_info(raw_city)
    city_slug = info[0] if info else 'bc'

    key = (city_slug, cat_slug, biz_slug)
    used_slugs[key] = used_slugs.get(key, 0) + 1
    final_slug = biz_slug if used_slugs[key] == 1 else f'{biz_slug}-{used_slugs[key]}'

    return f'{BASE}/{city_slug}/{cat_slug}/{final_slug}/'

def main():
    print(f'Opening {XLSX} ...')
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    print(f'  Sheet: "{ws.title}", {ws.max_row} rows, {ws.max_column} cols')

    # Read header row → map name → column index
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[v] = c

    # Set the new header in column N
    header_cell = ws.cell(row=1, column=URL_COL, value='Demo URL')
    header_cell.font = Font(bold=True, name='Calibri', size=11)

    # Widen column N so the long URL is readable when not hyperlinked
    ws.column_dimensions[openpyxl.utils.get_column_letter(URL_COL)].width = 75

    # Iterate data rows, build URLs (replay generator's collision logic across the sheet order)
    used_slugs = {}
    added = 0
    skipped = 0
    for r in range(2, ws.max_row + 1):
        row_dict = {h: ws.cell(row=r, column=c).value for h, c in headers.items()}
        url = build_url(row_dict, used_slugs)
        if url:
            cell = ws.cell(row=r, column=URL_COL, value=url)
            cell.hyperlink = url
            cell.font = LINK_FONT
            added += 1
        else:
            skipped += 1

    # Save (overwrites original)
    wb.save(XLSX)
    print(f'  Added {added} hyperlinked URLs ({skipped} rows had no matching demo)')
    print(f'  Saved back to: {XLSX}')

if __name__ == '__main__':
    main()
